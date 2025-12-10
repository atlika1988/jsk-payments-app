import streamlit as st
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from io import BytesIO
from payments_import import (
    import_statement,
    save_payments_to_db,
    Base,
    Apartment,
)


# ======================================================
# БАЗА ДАННЫХ
# ======================================================

DATABASE_URL = "sqlite:///jsk.db"
engine = create_engine(DATABASE_URL, echo=False)
SessionLocal = sessionmaker(bind=engine)
Base.metadata.create_all(engine)


def payments_to_dataframe(payments):
    """Табличное представление ParsedPayment."""
    return pd.DataFrame([
        {
            "Дата": p.date.strftime("%Y-%m-%d"),
            "Сумма": float(p.amount),
            "Описание": p.description,
            "Отправитель": p.sender_info,
            "Авто определение квартиры": p.guessed_apartment_number,
            "Выбранная квартира": p.apartment_id,
        }
        for p in payments
    ])


def load_apartment_map(session):
    """Список квартир в виде словаря 'Кв 61' → apartment.id."""
    apts = session.query(Apartment).order_by(Apartment.number).all()
    return {f"Кв {a.number}": a.id for a in apts}


# ======================================================
# UI
# ======================================================

st.title("📄 Импорт выписки СберБизнес — ЖСК 'Руслан'")

session = SessionLocal()
apt_map = load_apartment_map(session)

st.header("1. Загрузка файла выписки")
uploaded = st.file_uploader("Выберите файл .xlsx", type=["xlsx"])

if uploaded:
    temp_path = "uploaded_file.xlsx"
    with open(temp_path, "wb") as f:
        f.write(uploaded.read())

    st.success("Файл загружен. Обрабатываю…")

    matched, unmatched = import_statement(temp_path, session)

    st.subheader("2. Автоматически распознанные платежи")
    df_matched = payments_to_dataframe(matched)
    st.dataframe(df_matched, use_container_width=True)

    st.subheader("3. Платежи, требующие ручного сопоставления")

    # Заголовок таблицы
    cols = st.columns([1, 1, 3, 1, 1])
    cols[0].markdown("**Дата**")
    cols[1].markdown("**Сумма**")
    cols[2].markdown("**Описание**")
    cols[3].markdown("**Авто**")
    cols[4].markdown("**Квартира**")

    # строки таблицы
    for idx, p in enumerate(unmatched):
        row = st.columns([1, 1, 3, 1, 1])

        row[0].write(p.date.strftime("%Y-%m-%d"))
        row[1].write(float(p.amount))
        row[2].write(p.description)
        row[3].write(p.guessed_apartment_number)

        choice = row[4].selectbox(
            "",
            ["Не выбрано"] + list(apt_map.keys()),
            key=f"apt_choice_{idx}"
        )

        if choice != "Не выбрано":
            p.apartment_id = apt_map[choice]

    # ======================================================
    # КНОПКА ПРОВЕСТИ
    # ======================================================
    if st.button("📌 Провести платежи"):

        final_matched = matched + [p for p in unmatched if p.apartment_id]
        final_unmatched = [p for p in unmatched if not p.apartment_id]

        save_payments_to_db(session, final_matched, final_unmatched)

        st.success("Платежи успешно сохранены!")

    # ======================================================
    # ОТЧЁТ EXCEL — автофильтры, стиль, заморозка
    # ======================================================

    def autofit_columns(worksheet, dataframe):
        """Автоматически подгоняет ширину колонок под содержимое."""
        for i, col in enumerate(dataframe.columns):
            maxlen = max(
                [len(str(col))] + [len(str(v)) for v in dataframe[col].values]
            )
            worksheet.column_dimensions[worksheet.cell(row=1, column=i + 1).column_letter].width = maxlen + 2

    def create_report():
        """Создаёт Excel-файл с оформлением."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        output = BytesIO()
        wb = Workbook()

        # --------------- Лист 1: Распознанные ----------------
        ws1 = wb.active
        ws1.title = "Распознанные"

        df1 = payments_to_dataframe(matched)

        # Загружаем данные
        for r_idx, row in enumerate(df1.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws1.cell(row=r_idx, column=c_idx, value=val)

        # Заголовки
        for c_idx, col_name in enumerate(df1.columns, start=1):
            cell = ws1.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # автофильтр
        ws1.auto_filter.ref = ws1.dimensions

        # заморозка верхней строки
        ws1.freeze_panes = "A2"

        autofit_columns(ws1, df1)

        # --------------- Лист 2: Нераспознанные ----------------
        ws2 = wb.create_sheet("Нераспознанные")

        df2 = payments_to_dataframe(unmatched)

        for r_idx, row in enumerate(df2.values, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=val)

        for c_idx, col_name in enumerate(df2.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFEEDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        ws2.auto_filter.ref = ws2.dimensions
        ws2.freeze_panes = "A2"

        autofit_columns(ws2, df2)

        wb.save(output)
        output.seek(0)
        return output

    st.download_button(
        label="📥 Скачать отчёт",
        data=create_report(),
        file_name="Отчёт_выписки.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
